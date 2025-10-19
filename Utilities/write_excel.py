import os
import openpyxl
from openpyxl.styles import Font


def write_to_excel(sheet_name,result,filepath):
    if os.path.exists(filepath):
        wb=openpyxl.load_workbook(filepath)
        sheet=wb.create_sheet(sheet_name)
    else:
        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet.title=sheet_name
    sheet.cell(1,1,value='Car Model').font=Font(bold=True)
    sheet.cell(1,2,value='Price').font=Font(bold=True)
    for i in range(len(result)):
        sheet.append(result[i])
    wb.save(filepath)
