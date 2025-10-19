import openpyxl
import pytest
def readExcel(file_path):
    wb=openpyxl.load_workbook(file_path)
    ws=wb.active
    lst=[]
    for i in range(2,ws.max_row+1):
        l=[]
        for j in range(1,ws.max_column+1):
            l.append(ws.cell(i,j).value)
        lst.append(l)
    return lst
